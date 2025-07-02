import os
import openpyxl
from openpyxl.formatting.rule import ColorScaleRule

files = [f for f in os.listdir() if os.path.isfile(f) and f[-5:]==".xlsx"]
print(files)

excel = ""
while True:
    try:
        excel = int(input(f"Seleccione un archivo: 0 al {len(files)-1}: "))
        if excel < len(files):
            break
    except:
        pass
    print("Valor incorrecto")


wb = openpyxl.load_workbook(files[excel])
ws = wb.active

color_rule = ColorScaleRule(start_type='min', start_color='FFFFFF',
                            mid_type='percentile', mid_value=50, mid_color='FFFF00',
                            end_type='max', end_color='FF0000')

row = 2
while row <= ws.max_row:
    if ws[f'A{row}'].value == "":
        row += 1
        continue

    cell_range = f'B{row}:I{row+7}'
    ws.conditional_formatting.add(cell_range, color_rule)
    row += 9
new_file = files[excel][:-5]+"_colored.xlsx"
wb.save(new_file)

